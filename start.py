import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
lst = [1,2,3,4,5]

for row in range(2,sheet.max_row+1):
    cel = sheet.cell(row, 4)
    cel.value = lst[row]

wb.save("new.xlsx")

